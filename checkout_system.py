from openpyxl import *
from datetime import datetime
# Check to see if the excel file exists (program was already run)
try:
    wb = load_workbook(filename= 'tool_checkout_system.xlsx')
except:
    # Create the workbook if it doesnt exist
    wb = Workbook()
    toolCheckoutLogSheet = wb.create_sheet("tool_checkout_log")
    employeesSheet = wb.create_sheet("employees")
    toolsSheet = wb.create_sheet("tools")
    toolCheckoutLogSheet['A1'] = "Sign Out Employee"
    toolCheckoutLogSheet['B1'] = 'Tool'
    toolCheckoutLogSheet['C1'] = 'Sign Out Time'
    toolCheckoutLogSheet['D1'] = 'Sign In Time'
    toolCheckoutLogSheet['E1'] = 'Sign In Employee'


    employeesSheet['A1'] = 'Employee Number'
    employeesSheet['B1'] = 'Employee Name'

    toolsSheet['A1'] = 'Tool Number'
    toolsSheet['B1'] = 'Tool Name'
    toolsSheet['C1'] = 'Status'
    wb.save('tool_checkout_system.xlsx')

# Input: a sheet
# Output: An array of arrays containing the number followed by the name 
def getSheetData(sheet):
    data = []
    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column): 
        row_data = []
        for cell in row: 
            row_data.append(cell.value)
        data.append(row_data)
    return data


# Input: a sheet
# Output: a hashmap with keys as the numbers and values as the name
def initializeHashMap(data,map):
    for array in data:
        number = array[0]
        name = array[1]
        map[number] = name
    
# Input: employeenumber and toolnumber
# Output: updates sheets and activetool list
def signOutTool(employeeNumber,toolNumber):
    if toolNumber in activeTools:
        print("\n*** ERROR: This tool is already signed out. ***\n")
        return
    else:
        toolName = tools.get(toolNumber);
        employeeName = employees.get(employeeNumber)
        now = datetime.now()
        currentTime = now.strftime("%m/%d/%Y %H:%M")
        dataToAppend = [employeeName,toolName,currentTime]
        tool_checkout_log_sheet.append(dataToAppend)
        setToolAsActive(toolNumber)
        print(f"\n{toolName} SIGNED OUT BY {employeeName} AT {currentTime}\n")
        wb.save('tool_checkout_system.xlsx')

# Input: employeenumber and toolnumber
# Output: updates sheets and activetool list
        
def signInTool(employeeNumber,toolNumber):
    # iterate through the tools column
    #   if we find the tool name, 
    #       update sign in column and remove tool from active tools
    # keep track of the row we're on

    if toolNumber not in activeTools:
        print("\n*** ERROR: This tool is not signed out. ***\n")
        return



    currentRow = 1 
    for row in tool_checkout_log_sheet.iter_rows(min_row=2, min_col=1, max_row=tool_checkout_log_sheet.max_row, max_col=tool_checkout_log_sheet.max_column): 
        currentRow += 1
        for cell in row: 
            signInRowToCheck = 'D' + str(currentRow)
            # checking the rows for the tool name and making sure its not a tool that was previously signed out by checking the signed out column
            if cell.value == tools.get(toolNumber) and tool_checkout_log_sheet[signInRowToCheck].value is None:
                now = datetime.now()
                currentTime = now.strftime("%m/%d/%Y %H:%M")
                tool_checkout_log_sheet[signInRowToCheck] = currentTime
                tool_checkout_log_sheet[signInRowToCheck] = currentTime

                cellToUpdateSignInEmployee = 'E' + str(currentRow)
                tool_checkout_log_sheet[cellToUpdateSignInEmployee] = employees.get(employeeNumber)

                print(f"\n{tools.get(toolNumber)} SIGNED IN BY {employees.get(employeeNumber)} AT {currentTime}\n")
                removeToolAsActive(toolNumber)
                # activeTools.remove(toolNumber)
                wb.save('tool_checkout_system.xlsx')

# Input: none
# Output: creates active tools list
def getToolStatuses():
    # reset active tools every time we check
    activeTools.clear()
    currentRow = 1
    for col in tool_sheet.iter_cols(min_row=2,min_col=3,max_col=3,max_row=tool_sheet.max_row):
        for cell in col:
            currentRow += 1
            if cell.value == 'Active':
                toAppend = 'A' + str(currentRow)
                activeTools.append(tool_sheet[toAppend].value)


# Input: toolnumber
# Output: updates tools sheet to list active tools
def setToolAsActive(toolNumber):
    currentRow = 1 
    for row in tool_sheet.iter_rows(min_row=2, min_col=1, max_row=tool_sheet.max_row, max_col=tool_sheet.max_column): 
        currentRow += 1
        for cell in row: 
            if cell.value == toolNumber:
                statusRowToUpdate = 'C' + str(currentRow)
                tool_sheet[statusRowToUpdate] = 'Active'
                wb.save('tool_checkout_system.xlsx')

# Input: toolnumber
# Output: updates tools sheet to remove active tools
def removeToolAsActive(toolNumber):
    currentRow = 1 
    for row in tool_sheet.iter_rows(min_row=2, min_col=1, max_row=tool_sheet.max_row, max_col=tool_sheet.max_column): 
        currentRow += 1
        for cell in row: 
            if cell.value == toolNumber:
                statusRowToUpdate = 'C' + str(currentRow)
                tool_sheet[statusRowToUpdate] = None
                wb.save('tool_checkout_system.xlsx')

# Dictionaries to store and retrieve data
    
# Key : barcode number
# Value: tool name
tools = {}

# Key : barcode number
# Value: employee name
employees = {}

# Iterate over the sheets and extract data
tool_sheet = wb['tools']
employee_sheet = wb['employees']
tool_checkout_log_sheet = wb['tool_checkout_log']
toolsData = getSheetData(tool_sheet)
employeesData = getSheetData(employee_sheet)

# Initialize the hashmaps
initializeHashMap(toolsData,tools)
initializeHashMap(employeesData,employees)

# Array to store active tools
activeTools = []
# print(tools)
# print(employees)
while True:
    getToolStatuses()
    print("\nWelcome to the tool checkout system")
    print("------------------------------------")
    # Collect and Validate input
    isValidInput = False
    while not isValidInput:
        action = input("Enter 1 to SIGN OUT, 2 to SIGN IN, or 'q' to QUIT: ")
        if action not in ['1','2','q']:
            print("\n*** ERROR: Invalid Input. ***\n")
        else:
            isValidInput = True
            if action == 'q':
                wb.save('tool_checkout_system.xlsx')
                wb.close()
                quit()
    # Get employee barcode number
            
    employeeNumber = int(input("Scan your employee badge: "))
    toolNumber = int(input("Scan the tool barcode: "))

    # make sure the inputs are valid
    if employeeNumber not in employees.keys():
        print("\n*** ERROR: Invalid employee number. ***\n")
        continue
    elif toolNumber not in tools.keys():
        print("\n*** ERROR: Invalid tool number. ***\n")
        continue


    
    match action:
        case "1":
            signOutTool(employeeNumber,toolNumber)
        case "2": 
            signInTool(employeeNumber,toolNumber)


    

    